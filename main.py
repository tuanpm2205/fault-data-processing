import pandas as pd
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== 1. NHẬP THÔNG TIN =====
# Nhập đường dẫn file .xls (tự bỏ dấu " hoặc ' nếu copy từ Explorer)
print(r"Ví dụ: C:\Users\phamm\Documents\Python\Fault_20260211081942-1.xls")
duong_dan_file = input("Nhập đường dẫn file đầu vào (.xls): ").strip('"').strip("'")

# Kiểm tra file tồn tại trước khi đọc
if not os.path.exists(duong_dan_file):
    print("Lỗi: Không tìm thấy file.")
    exit()

# Nhập ngày cần lọc
nam = input("Nam : ")
thang = input("Thang : ").zfill(2)  # đảm bảo 2 chữ số
ngay = input("Ngay : ").zfill(2)

# Tên file xuất
ten_file_xuat = input("Ten file xuat: ") + ".xlsx"


# ===== 2. ĐỌC FILE XLS =====
print("Đang đọc dữ liệu...")
try:
    df = pd.read_excel(
        duong_dan_file,
        sheet_name=0,     # luôn đọc sheet đầu tiên
        engine='xlrd'     # bắt buộc cho file .xls cũ
    )
except Exception as e:
    print(f"Lỗi khi đọc file: {e}")
    exit()


# ===== 3. CẤU HÌNH CỘT SỬ DỤNG =====
# Đổi lại nếu file nguồn đổi tên cột
cot_can_loc = "Plant name"        # dùng để tạo sheet
cot_chia_bang_con = "Device name" # chia bảng con trong mỗi sheet
ten_cot_ngay_thang = "Recovery time"


# ===== 4. LỌC THEO NGÀY =====
gia_tri_chuan = nam + "-" + thang + "-" + ngay

# Kiểm tra tồn tại cột ngày
if ten_cot_ngay_thang not in df.columns:
    print(f"Lỗi: Không tìm thấy cột '{ten_cot_ngay_thang}'.")
    exit()

# Chuyển sang datetime và lọc theo ngày
df[ten_cot_ngay_thang] = pd.to_datetime(df[ten_cot_ngay_thang], errors='coerce')
df['Ngay_de_so_sanh'] = df[ten_cot_ngay_thang].dt.strftime('%Y-%m-%d')
df = df[df['Ngay_de_so_sanh'] == gia_tri_chuan]
df = df.drop(columns=['Ngay_de_so_sanh'])

if df.empty:
    print(f"Không có dữ liệu ngày {gia_tri_chuan}.")
    exit()

print(f"Đang xuất file: {ten_file_xuat}")


# ===== 5. CẤU HÌNH STYLE EXCEL =====
font_arial = Font(name='Arial', size=11, bold=False)
fill_blue = PatternFill(start_color="87CEEB", fill_type="solid")
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ===== 6. GHI FILE VÀ FORMAT =====
with pd.ExcelWriter(ten_file_xuat, engine='openpyxl') as writer:

    # Tạo 1 sheet cho mỗi Plant
    for gia_tri, nhom_du_lieu in df.groupby(cot_can_loc):

        # Chuẩn hóa tên sheet (Excel giới hạn 31 ký tự)
        ten_sheet = str(gia_tri)[:31].replace("/", "-").replace(":", "").replace("?", "")
        current_row = 0

        # Trong mỗi sheet, chia tiếp theo Device
        for ten_bang_con, data_bang_con in nhom_du_lieu.groupby(cot_chia_bang_con):

            # Ghi bảng
            data_bang_con.to_excel(
                writer,
                sheet_name=ten_sheet,
                startrow=current_row,
                index=False
            )

            worksheet = writer.sheets[ten_sheet]
            row_header_idx = current_row + 1
            num_cols = len(data_bang_con.columns)

            # --- Format header ---
            worksheet.row_dimensions[row_header_idx].height = 25.7

            for col_idx in range(1, num_cols + 1):
                cell = worksheet.cell(row=row_header_idx, column=col_idx)
                cell.font = font_arial
                cell.fill = fill_blue
                cell.alignment = align_center
                cell.border = thin_border

                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 23.22

            # --- Format data ---
            for r_idx in range(row_header_idx + 1,
                               row_header_idx + 1 + len(data_bang_con)):
                worksheet.row_dimensions[r_idx].height = 25.7
                for c_idx in range(1, num_cols + 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    cell.font = font_arial
                    cell.alignment = align_center
                    cell.border = thin_border

            # Cách 1 dòng trống giữa các bảng con
            current_row = current_row + 1 + len(data_bang_con) + 1

print(f"Xong. File đã tạo: {ten_file_xuat}")
